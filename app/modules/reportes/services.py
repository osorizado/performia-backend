"""
Servicios de reportes y notificaciones
"""
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, case
from typing import List, Optional
from datetime import datetime
import json
import os
from pathlib import Path

# Librerías para generación de PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER

# Librerías para generación de Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.modules.reportes.models import Reporte, Notificacion
from app.modules.reportes.schemas import (
    ReporteCreate, 
    NotificacionCreate,
    EstadisticasGenerales,
    EstadisticasCompetencias,
    TopPerformer,
    AreaRanking,
    FiltrosReporte
)
from app.modules.evaluaciones.models import Evaluacion, Resultado
from app.modules.formularios.models import Formulario, Pregunta
from app.modules.users.models import Usuario


# ========================================
# SERVICIOS DE NOTIFICACIONES
# ========================================

def create_notificacion(db: Session, notif: NotificacionCreate) -> Notificacion:
    """Crea una notificación"""
    db_notif = Notificacion(**notif.model_dump())
    db.add(db_notif)
    db.commit()
    db.refresh(db_notif)
    return db_notif


def get_mis_notificaciones(db: Session, usuario_id: int) -> List[Notificacion]:
    """Obtiene las notificaciones de un usuario"""
    return db.query(Notificacion).filter(
        Notificacion.id_usuario == usuario_id
    ).order_by(Notificacion.fecha_envio.desc()).all()


def marcar_notificacion_leida(db: Session, notif_id: int) -> Notificacion:
    """Marca una notificación como leída"""
    notif = db.query(Notificacion).filter(
        Notificacion.id_notificacion == notif_id
    ).first()
    
    if notif:
        notif.leida = True
        db.commit()
        db.refresh(notif)
    
    return notif


# ========================================
# SERVICIOS DE ESTADÍSTICAS
# ========================================

def get_estadisticas_generales(db: Session, periodo: Optional[str] = None) -> EstadisticasGenerales:
    """
    Calcula las estadísticas generales del sistema.
    
    Args:
        periodo: Filtro de período (opcional)
        
    Returns:
        EstadisticasGenerales con métricas del sistema
    """
    # Query base de evaluaciones
    query_evaluaciones = db.query(Evaluacion)
    
    # Aplicar filtro de período si existe
    if periodo:
        query_evaluaciones = query_evaluaciones.filter(Evaluacion.periodo == periodo)
    
    # Total de evaluaciones completas
    evaluaciones_completas = query_evaluaciones.filter(
        Evaluacion.estado == 'Completada'
    ).count()
    
    # Total de evaluaciones pendientes
    evaluaciones_pendientes = query_evaluaciones.filter(
        Evaluacion.estado.in_(['Pendiente', 'En Curso'])
    ).count()
    
    # Total de evaluaciones
    total_evaluaciones = evaluaciones_completas + evaluaciones_pendientes
    
    # Tasa de completitud
    tasa_completitud = (evaluaciones_completas / total_evaluaciones * 100) if total_evaluaciones > 0 else 0
    
    # Promedio general (de todos los puntajes en evaluaciones completas)
    promedio_query = db.query(
        func.avg(Resultado.puntaje)
    ).join(
        Evaluacion, Resultado.id_evaluacion == Evaluacion.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None)
    )
    
    if periodo:
        promedio_query = promedio_query.filter(Evaluacion.periodo == periodo)
    
    promedio_general = promedio_query.scalar()
    promedio_general = float(promedio_general) if promedio_general else 0.0
    
    # Top performers: colaboradores con promedio >= 4.5
    # Subconsulta para calcular el promedio por evaluado
    subq = db.query(
        Evaluacion.id_evaluado,
        func.avg(Resultado.puntaje).label('promedio')
    ).join(
        Resultado, Evaluacion.id_evaluacion == Resultado.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None)
    )
    
    if periodo:
        subq = subq.filter(Evaluacion.periodo == periodo)
    
    subq = subq.group_by(Evaluacion.id_evaluado).subquery()
    
    top_performers = db.query(subq).filter(subq.c.promedio >= 4.5).count()
    
    # Total de colaboradores únicos evaluados
    total_colaboradores_query = db.query(
        func.count(func.distinct(Evaluacion.id_evaluado))
    ).filter(
        Evaluacion.estado == 'Completada'
    )
    
    if periodo:
        total_colaboradores_query = total_colaboradores_query.filter(Evaluacion.periodo == periodo)
    
    total_colaboradores = total_colaboradores_query.scalar() or 0
    
    # Total de evaluadores únicos
    total_evaluadores_query = db.query(
        func.count(func.distinct(Evaluacion.id_evaluador))
    ).filter(
        Evaluacion.estado == 'Completada'
    )
    
    if periodo:
        total_evaluadores_query = total_evaluadores_query.filter(Evaluacion.periodo == periodo)
    
    total_evaluadores = total_evaluadores_query.scalar() or 0
    
    return EstadisticasGenerales(
        promedio_general=round(promedio_general, 2),
        evaluaciones_completas=evaluaciones_completas,
        evaluaciones_pendientes=evaluaciones_pendientes,
        tasa_completitud=round(tasa_completitud, 1),
        top_performers=top_performers,
        total_colaboradores=total_colaboradores,
        total_evaluadores=total_evaluadores
    )


def get_estadisticas_competencias(
    db: Session, 
    periodo: Optional[str] = None,
    area: Optional[str] = None
) -> List[EstadisticasCompetencias]:
    """
    Calcula el promedio de puntaje por competencia.
    Útil para gráficos de radar.
    """
    # Query base
    query = db.query(
        Pregunta.competencia,
        func.avg(Resultado.puntaje).label('promedio'),
        func.count(Resultado.id_resultado).label('cantidad')
    ).join(
        Resultado, Pregunta.id_pregunta == Resultado.id_pregunta
    ).join(
        Evaluacion, Resultado.id_evaluacion == Evaluacion.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None),
        Pregunta.competencia.isnot(None)
    )
    
    # Filtro por período
    if periodo:
        query = query.filter(Evaluacion.periodo == periodo)
    
    # Filtro por área (si se especifica)
    if area:
        query = query.join(
            Usuario, Evaluacion.id_evaluado == Usuario.id_usuario
        ).filter(Usuario.area == area)
    
    # Agrupar por competencia
    query = query.group_by(Pregunta.competencia)
    
    resultados = query.all()
    
    return [
        EstadisticasCompetencias(
            competencia=r.competencia,
            promedio=round(float(r.promedio), 2),
            cantidad_evaluaciones=r.cantidad
        )
        for r in resultados
    ]


def get_distribucion_calificaciones(db: Session, periodo: Optional[str] = None) -> dict:
    """
    Obtiene la distribución de colaboradores por rango de calificación.
    Retorna un diccionario con rangos como claves.
    """
    # Subconsulta: promedio por colaborador
    subq = db.query(
        Evaluacion.id_evaluado,
        func.avg(Resultado.puntaje).label('promedio')
    ).join(
        Resultado, Evaluacion.id_evaluacion == Resultado.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None)
    )
    
    if periodo:
        subq = subq.filter(Evaluacion.periodo == periodo)
    
    subq = subq.group_by(Evaluacion.id_evaluado).subquery()
    
    # Contar colaboradores por rango
    rango_1 = db.query(subq).filter(and_(subq.c.promedio >= 0, subq.c.promedio <= 2.0)).count()
    rango_2 = db.query(subq).filter(and_(subq.c.promedio > 2.0, subq.c.promedio <= 3.0)).count()
    rango_3 = db.query(subq).filter(and_(subq.c.promedio > 3.0, subq.c.promedio <= 4.0)).count()
    rango_4 = db.query(subq).filter(and_(subq.c.promedio > 4.0, subq.c.promedio <= 5.0)).count()
    
    return {
        "1.0-2.0": rango_1,
        "2.1-3.0": rango_2,
        "3.1-4.0": rango_3,
        "4.1-5.0": rango_4
    }


def get_top_performers(
    db: Session, 
    limite: int = 10,
    periodo: Optional[str] = None
) -> List[TopPerformer]:
    """
    Obtiene el ranking de los mejores colaboradores.
    """
    # Subconsulta: promedio y cantidad de evaluaciones por colaborador
    subq = db.query(
        Evaluacion.id_evaluado,
        func.avg(Resultado.puntaje).label('promedio'),
        func.count(func.distinct(Evaluacion.id_evaluacion)).label('cant_evaluaciones')
    ).join(
        Resultado, Evaluacion.id_evaluacion == Resultado.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None)
    )
    
    if periodo:
        subq = subq.filter(Evaluacion.periodo == periodo)
    
    subq = subq.group_by(Evaluacion.id_evaluado).subquery()
    
    # Unir con datos del usuario y ordenar
    resultados = db.query(
        Usuario.id_usuario,
        Usuario.nombre,
        Usuario.apellido,
        Usuario.area,
        Usuario.cargo,
        subq.c.promedio,
        subq.c.cant_evaluaciones
    ).join(
        subq, Usuario.id_usuario == subq.c.id_evaluado
    ).order_by(
        subq.c.promedio.desc()
    ).limit(limite).all()
    
    return [
        TopPerformer(
            id_usuario=r.id_usuario,
            nombre=r.nombre or "",
            apellido=r.apellido or "",
            area=r.area,
            cargo=r.cargo,
            promedio=round(float(r.promedio), 2),
            evaluaciones_completas=r.cant_evaluaciones
        )
        for r in resultados
    ]


def get_areas_ranking(db: Session) -> List[AreaRanking]:
    """
    Obtiene el ranking de áreas por desempeño promedio.
    """
    # Subconsulta: promedio por colaborador
    subq_promedio = db.query(
        Evaluacion.id_evaluado,
        func.avg(Resultado.puntaje).label('promedio')
    ).join(
        Resultado, Evaluacion.id_evaluacion == Resultado.id_evaluacion
    ).filter(
        Evaluacion.estado == 'Completada',
        Resultado.puntaje.isnot(None)
    ).group_by(
        Evaluacion.id_evaluado
    ).subquery()
    
    # Agrupar por área
    resultados = db.query(
        Usuario.area,
        func.avg(subq_promedio.c.promedio).label('promedio_area'),
        func.count(func.distinct(Usuario.id_usuario)).label('total_colaboradores'),
        func.count(func.distinct(Evaluacion.id_evaluacion)).label('evaluaciones_completas')
    ).join(
        subq_promedio, Usuario.id_usuario == subq_promedio.c.id_evaluado
    ).join(
        Evaluacion, Usuario.id_usuario == Evaluacion.id_evaluado
    ).filter(
        Usuario.area.isnot(None),
        Evaluacion.estado == 'Completada'
    ).group_by(
        Usuario.area
    ).order_by(
        func.avg(subq_promedio.c.promedio).desc()
    ).all()
    
    return [
        AreaRanking(
            area=r.area,
            promedio=round(float(r.promedio_area), 2),
            total_colaboradores=r.total_colaboradores,
            evaluaciones_completas=r.evaluaciones_completas
        )
        for r in resultados
    ]


# ========================================
# FUNCIONES AUXILIARES PARA OBTENER DATOS
# ========================================

def obtener_datos_reporte_global(db: Session, periodo: Optional[str] = None):
    """Obtiene todos los datos necesarios para el reporte global"""
    return {
        "estadisticas_generales": get_estadisticas_generales(db, periodo),
        "estadisticas_competencias": get_estadisticas_competencias(db, periodo),
        "top_performers": get_top_performers(db, limite=10, periodo=periodo),
        "distribucion_calificaciones": get_distribucion_calificaciones(db, periodo),
        "areas_ranking": get_areas_ranking(db)
    }


def generar_pdf(datos: dict, ruta_archivo: str, config: dict):
    """Genera un archivo PDF profesional con los datos del reporte"""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors as pdf_colors
    
    doc = SimpleDocTemplate(
        ruta_archivo, 
        pagesize=letter,
        rightMargin=50, 
        leftMargin=50,
        topMargin=50, 
        bottomMargin=50
    )
    story = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados mejorados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#4a5568'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=18,
        textColor=colors.HexColor('#2d3748'),
        spaceAfter=15,
        spaceBefore=20,
        fontName='Helvetica-Bold',
        borderPadding=10,
        leftIndent=0
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#718096'),
        alignment=TA_CENTER
    )
    
    # Encabezado del documento
    story.append(Paragraph("PERFORMIA", title_style))
    story.append(Paragraph(f"Reporte de Desempeño - {config['tipo_reporte']}", subtitle_style))
    
    # Línea separadora
    from reportlab.platypus import HRFlowable
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#4299e1'), spaceBefore=10, spaceAfter=20))
    
    # Información del reporte
    fecha_texto = f"<b>Generado:</b> {datetime.now().strftime('%d de %B de %Y a las %H:%M')}"
    if config.get('periodo'):
        fecha_texto += f" | <b>Período:</b> {config['periodo']}"
    story.append(Paragraph(fecha_texto, info_style))
    story.append(Spacer(1, 0.4*inch))
    
    # RESUMEN EJECUTIVO
    story.append(Paragraph("📊 Resumen Ejecutivo", heading_style))
    stats = datos["estadisticas_generales"]
    
    # Tabla de resumen con diseño mejorado
    resumen_data = [
        ['MÉTRICA', 'VALOR', 'ESTADO'],
        ['Promedio General', f"{stats.promedio_general:.2f}/5.0", 
         '🟢 Excelente' if stats.promedio_general >= 4.5 else '🟡 Bueno' if stats.promedio_general >= 3.5 else '🔴 Mejorar'],
        ['Evaluaciones Completadas', str(stats.evaluaciones_completas), '✓'],
        ['Evaluaciones Pendientes', str(stats.evaluaciones_pendientes), 
         '⚠️' if stats.evaluaciones_pendientes > 0 else '✓'],
        ['Tasa de Completitud', f"{stats.tasa_completitud:.1f}%",
         '🟢' if stats.tasa_completitud >= 80 else '🟡' if stats.tasa_completitud >= 60 else '🔴'],
        ['Top Performers (≥4.5)', str(stats.top_performers), '⭐'],
        ['Total Colaboradores', str(stats.total_colaboradores), '👥'],
        ['Total Evaluadores', str(stats.total_evaluadores), '👤']
    ]
    
    resumen_table = Table(resumen_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    resumen_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
        ('TOPPADDING', (0, 0), (-1, 0), 15),
        
        # Cuerpo
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f7fafc')),
        ('BACKGROUND', (1, 1), (1, -1), colors.white),
        ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#edf2f7')),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ('LEFTPADDING', (0, 1), (0, -1), 15),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e0')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#4299e1')),
        
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 0.4*inch))
    
    # TOP 10 PERFORMERS
    if datos.get("top_performers"):
        story.append(Paragraph("🏆 Top 10 Mejores Colaboradores", heading_style))
        
        top_data = [['#', 'COLABORADOR', 'ÁREA', 'CARGO', 'PROMEDIO', 'EVAL.']]
        
        for idx, performer in enumerate(datos["top_performers"][:10], 1):
            medalla = '🥇' if idx == 1 else '🥈' if idx == 2 else '🥉' if idx == 3 else str(idx)
            promedio = performer['promedio']
            promedio_icon = '⭐⭐⭐⭐⭐' if promedio >= 4.8 else '⭐⭐⭐⭐' if promedio >= 4.0 else '⭐⭐⭐'
            
            top_data.append([
                medalla,
                f"{performer['nombre']} {performer['apellido']}",
                performer['area'] or 'N/A',
                performer['cargo'] or 'N/A',
                f"{promedio:.2f}\n{promedio_icon}",
                str(performer['cant_evaluaciones'])
            ])
        
        top_table = Table(top_data, colWidths=[0.4*inch, 2*inch, 1.2*inch, 1.2*inch, 1*inch, 0.6*inch])
        top_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Cuerpo
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (1, 1), (1, -1), 10),
            
            # Top 3 destacado
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#fef3c7')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fde8e8')),
            
            # Resto alternado
            ('ROWBACKGROUNDS', (0, 4), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#48bb78')),
        ]))
        story.append(top_table)
        story.append(Spacer(1, 0.3*inch))
    
    # COMPETENCIAS (Nueva página)
    if datos.get("estadisticas_competencias"):
        story.append(PageBreak())
        story.append(Paragraph("📈 Análisis por Competencias", heading_style))
        
        comp_data = [['COMPETENCIA', 'PROMEDIO', 'EVALUACIONES', 'NIVEL']]
        for comp in datos["estadisticas_competencias"]:
            nivel = '🟢 Alto' if comp.promedio >= 4.5 else '🟡 Medio' if comp.promedio >= 3.5 else '🔴 Bajo'
            comp_data.append([
                comp.competencia,
                f"{comp.promedio:.2f}/5.0",
                str(comp.cantidad_evaluaciones),
                nivel
            ])
        
        comp_table = Table(comp_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        comp_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Cuerpo
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('LEFTPADDING', (0, 1), (0, -1), 15),
            
            # Bordes y colores
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#ed8936')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        story.append(comp_table)
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cbd5e0')))
    footer_text = f"<i>Reporte generado por PERFORMIA © {datetime.now().year} | Documento confidencial</i>"
    story.append(Paragraph(footer_text, info_style))
    
    doc.build(story)


def generar_excel(datos: dict, ruta_archivo: str, config: dict):
    """Genera un archivo Excel profesional con los datos del reporte"""
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ============================================
    # HOJA 1: RESUMEN EJECUTIVO
    # ============================================
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    
    # Título principal
    ws1['A1'] = "PERFORMIA - Sistema de Evaluación de Desempeño"
    ws1['A1'].font = Font(size=18, bold=True, color="1a365d")
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:E1')
    
    # Subtítulo
    ws1['A2'] = f"Reporte de Desempeño - {config['tipo_reporte']}"
    ws1['A2'].font = Font(size=14, color="4a5568")
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:E2')
    
    # Fecha
    ws1['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A3'].font = Font(size=10, color="718096", italic=True)
    ws1['A3'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A3:E3')
    
    # Estadísticas
    stats = datos["estadisticas_generales"]
    
    # Headers
    ws1['A5'] = "MÉTRICA"
    ws1['B5'] = "VALOR"
    ws1['C5'] = "ESTADO"
    
    for col in ['A5', 'B5', 'C5']:
        ws1[col].font = Font(bold=True, color="FFFFFF", size=12)
        ws1[col].fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
        ws1[col].alignment = Alignment(horizontal='center', vertical='center')
        ws1[col].border = Border(
            bottom=Side(style='medium', color='4299e1')
        )
    
    # Datos
    metricas_data = [
        ("Promedio General", f"{stats.promedio_general:.2f}/5.0", 
         "Excelente" if stats.promedio_general >= 4.5 else "Bueno" if stats.promedio_general >= 3.5 else "Mejorar"),
        ("Evaluaciones Completadas", stats.evaluaciones_completas, "✓"),
        ("Evaluaciones Pendientes", stats.evaluaciones_pendientes, 
         "⚠️" if stats.evaluaciones_pendientes > 0 else "✓"),
        ("Tasa de Completitud", f"{stats.tasa_completitud:.1f}%",
         "Alto" if stats.tasa_completitud >= 80 else "Medio" if stats.tasa_completitud >= 60 else "Bajo"),
        ("Top Performers (≥4.5)", stats.top_performers, "⭐"),
        ("Total Colaboradores", stats.total_colaboradores, "👥"),
        ("Total Evaluadores", stats.total_evaluadores, "👤"),
    ]
    
    border_thin = Border(
        left=Side(style='thin', color='cbd5e0'),
        right=Side(style='thin', color='cbd5e0'),
        top=Side(style='thin', color='cbd5e0'),
        bottom=Side(style='thin', color='cbd5e0')
    )
    
    for idx, (metrica, valor, estado) in enumerate(metricas_data, start=6):
        ws1[f'A{idx}'] = metrica
        ws1[f'B{idx}'] = valor
        ws1[f'C{idx}'] = estado
        
        # Formato alternado
        if idx % 2 == 0:
            fill_color = "f8f9fa"
        else:
            fill_color = "FFFFFF"
        
        ws1[f'A{idx}'].font = Font(size=11)
        ws1[f'A{idx}'].fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")
        ws1[f'A{idx}'].alignment = Alignment(horizontal='left', vertical='center')
        ws1[f'A{idx}'].border = border_thin
        
        ws1[f'B{idx}'].font = Font(size=11, bold=True)
        ws1[f'B{idx}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws1[f'B{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws1[f'B{idx}'].border = border_thin
        
        ws1[f'C{idx}'].font = Font(size=11)
        ws1[f'C{idx}'].fill = PatternFill(start_color="edf2f7", end_color="edf2f7", fill_type="solid")
        ws1[f'C{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws1[f'C{idx}'].border = border_thin
    
    # Ajustar columnas
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 25
    ws1.row_dimensions[5].height = 25
    
    # ============================================
    # HOJA 2: TOP PERFORMERS
    # ============================================
    if datos.get("top_performers"):
        ws2 = wb.create_sheet("Top Performers")
        
        # Título
        ws2['A1'] = "🏆 TOP 10 MEJORES COLABORADORES"
        ws2['A1'].font = Font(size=16, bold=True, color="1a365d")
        ws2['A1'].alignment = Alignment(horizontal='center')
        ws2.merge_cells('A1:G1')
        ws2.row_dimensions[1].height = 30
        
        # Headers
        headers = ['RANK', 'NOMBRE', 'APELLIDO', 'ÁREA', 'CARGO', 'PROMEDIO', 'EVAL.']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws2.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(bottom=Side(style='medium', color='48bb78'))
        
        ws2.row_dimensions[3].height = 25
        
        # Datos
        for idx, performer in enumerate(datos["top_performers"][:10], start=4):
            rank = idx - 3
            
            # Color especial para top 3
            if rank == 1:
                fill_color = "fef3c7"  # Dorado
            elif rank == 2:
                fill_color = "f3f4f6"  # Plata
            elif rank == 3:
                fill_color = "fde8e8"  # Bronce
            else:
                fill_color = "FFFFFF" if idx % 2 == 0 else "f8f9fa"
            
            cells_data = [
                (rank, 'center'),
                (performer['nombre'], 'left'),
                (performer['apellido'], 'left'),
                (performer['area'] or 'N/A', 'center'),
                (performer['cargo'] or 'N/A', 'center'),
                (round(performer['promedio'], 2), 'center'),
                (performer['cant_evaluaciones'], 'center')
            ]
            
            for col_idx, (value, align) in enumerate(cells_data, start=1):
                cell = ws2.cell(row=idx, column=col_idx)
                cell.value = value
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal=align, vertical='center')
                cell.border = border_thin
                
                if col_idx == 1:
                    cell.font = Font(bold=True, size=11)
                elif col_idx == 6:
                    cell.font = Font(bold=True, size=11, color="1a365d")
                else:
                    cell.font = Font(size=10)
            
            ws2.row_dimensions[idx].height = 22
        
        # Ajustar columnas
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 20
        ws2.column_dimensions['F'].width = 12
        ws2.column_dimensions['G'].width = 8
    
    # ============================================
    # HOJA 3: COMPETENCIAS
    # ============================================
    if datos.get("estadisticas_competencias"):
        ws3 = wb.create_sheet("Análisis de Competencias")
        
        # Título
        ws3['A1'] = "📈 ANÁLISIS POR COMPETENCIAS"
        ws3['A1'].font = Font(size=16, bold=True, color="1a365d")
        ws3['A1'].alignment = Alignment(horizontal='center')
        ws3.merge_cells('A1:D1')
        ws3.row_dimensions[1].height = 30
        
        # Headers
        headers = ['COMPETENCIA', 'PROMEDIO', 'EVALUACIONES', 'NIVEL']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws3.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(bottom=Side(style='medium', color='ed8936'))
        
        ws3.row_dimensions[3].height = 25
        
        # Datos
        for idx, comp in enumerate(datos["estadisticas_competencias"], start=4):
            nivel = "Alto" if comp.promedio >= 4.5 else "Medio" if comp.promedio >= 3.5 else "Bajo"
            nivel_color = "c6f6d5" if comp.promedio >= 4.5 else "fef9c3" if comp.promedio >= 3.5 else "fed7d7"
            
            fill_color = "FFFFFF" if idx % 2 == 0 else "f8f9fa"
            
            # Competencia
            cell_a = ws3.cell(row=idx, column=1)
            cell_a.value = comp.competencia
            cell_a.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell_a.alignment = Alignment(horizontal='left', vertical='center')
            cell_a.border = border_thin
            cell_a.font = Font(size=11)
            
            # Promedio
            cell_b = ws3.cell(row=idx, column=2)
            cell_b.value = f"{comp.promedio:.2f}/5.0"
            cell_b.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell_b.alignment = Alignment(horizontal='center', vertical='center')
            cell_b.border = border_thin
            cell_b.font = Font(size=11, bold=True, color="1a365d")
            
            # Evaluaciones
            cell_c = ws3.cell(row=idx, column=3)
            cell_c.value = comp.cantidad_evaluaciones
            cell_c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell_c.alignment = Alignment(horizontal='center', vertical='center')
            cell_c.border = border_thin
            cell_c.font = Font(size=11)
            
            # Nivel
            cell_d = ws3.cell(row=idx, column=4)
            cell_d.value = nivel
            cell_d.fill = PatternFill(start_color=nivel_color, end_color=nivel_color, fill_type="solid")
            cell_d.alignment = Alignment(horizontal='center', vertical='center')
            cell_d.border = border_thin
            cell_d.font = Font(size=11, bold=True)
            
            ws3.row_dimensions[idx].height = 22
        
        # Ajustar columnas
        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 18
        ws3.column_dimensions['D'].width = 15
    
    # Guardar archivo
    wb.save(ruta_archivo)


# ========================================
# SERVICIOS DE GENERACIÓN DE REPORTES
# ========================================

def generar_reporte(db: Session, filtros: FiltrosReporte, generado_por: int) -> Reporte:
    """
    Genera un reporte según los filtros especificados y crea el archivo físico (PDF o Excel).
    """
    try:
        reportes_dir = Path("reportes_generados")
        reportes_dir.mkdir(exist_ok=True)
        
        datos = obtener_datos_reporte_global(db, filtros.periodo)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        formato = filtros.formato
        tipo_clean = filtros.tipo_reporte.replace(' ', '_').replace('á', 'a').replace('ó', 'o')
        nombre_base = f"Reporte_{tipo_clean}_{timestamp}"
        
        if formato == "PDF":
            nombre_archivo = f"{nombre_base}.pdf"
            ruta_completa = reportes_dir / nombre_archivo
            generar_pdf(datos, str(ruta_completa), filtros.model_dump())
        elif formato == "Excel":
            nombre_archivo = f"{nombre_base}.xlsx"
            ruta_completa = reportes_dir / nombre_archivo
            generar_excel(datos, str(ruta_completa), filtros.model_dump())
        else:
            raise ValueError("Formato no válido")
        
        nuevo_reporte = Reporte(
            nombre_reporte=f"{filtros.tipo_reporte} - {timestamp}",
            tipo_reporte=filtros.tipo_reporte,
            periodo=filtros.periodo,
            parametros=json.dumps(filtros.model_dump()),
            ruta_archivo=str(ruta_completa),
            formato=formato,
            generado_por=generado_por
        )
        
        db.add(nuevo_reporte)
        db.commit()
        db.refresh(nuevo_reporte)
        
        return nuevo_reporte
        
    except Exception as e:
        db.rollback()
        print(f"Error al generar reporte: {str(e)}")
        raise


def get_reporte_by_id(db: Session, reporte_id: int) -> Optional[Reporte]:
    """Obtiene un reporte por ID"""
    return db.query(Reporte).filter(Reporte.id_reporte == reporte_id).first()


def get_historial_reportes(db: Session, limite: int = 20) -> List[Reporte]:
    """Obtiene el historial de reportes generados"""
    return db.query(Reporte).order_by(
        Reporte.fecha_generacion.desc()
    ).limit(limite).all()


def eliminar_reporte(db: Session, reporte_id: int) -> bool:
    """Elimina un reporte y su archivo físico"""
    reporte = db.query(Reporte).filter(Reporte.id_reporte == reporte_id).first()
    
    if reporte:
        if reporte.ruta_archivo and os.path.exists(reporte.ruta_archivo):
            try:
                os.remove(reporte.ruta_archivo)
            except Exception as e:
                print(f"Error al eliminar archivo: {str(e)}")
        
        db.delete(reporte)
        db.commit()
        return True
    
    return False